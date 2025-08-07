import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import os
import fitz
import pandas as pd
import openpyxl
from openai import OpenAI
from keys import open_ai_api_key

# Initialize OpenAI client
client = OpenAI(api_key=open_ai_api_key)

# ----------- Paragraph Extraction -----------

def extract_paragraphs(text, client):
    """
    Extracts paragraphs from text using OpenAI API.
    
    Uses AI to intelligently split raw text into meaningful paragraphs
    by streaming responses and parsing paragraph markers.
    
    Args:
        text (str): Raw text extracted from PDF
        client (OpenAI): Initialized OpenAI client for API calls
    
    Returns:
        list: List of extracted paragraph strings
    
    Example:
        >>> client = OpenAI(api_key="your-key")
        >>> text = "First paragraph. Second paragraph."
        >>> paragraphs = extract_paragraphs(text, client)
        >>> print(len(paragraphs))
        2
    """
    prompt = (
        "Split the following text into paragraphs. "
        "After each paragraph, output the marker <P>.\n\n"
        + text
    )
    paragraphs = []
    buffer = ""

    with client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[{"role": "user", "content": prompt}],
        stream=True,
        temperature=0
    ) as stream:
        for chunk in stream:
            if hasattr(chunk, "choices") and chunk.choices:
                delta = chunk.choices[0].delta
                if hasattr(delta, "content") and delta.content:
                    buffer += delta.content

    while "<P>" in buffer:
        para, buffer = buffer.split("<P>", 1)
        para = para.strip()
        if para:
            paragraphs.append(para)

    return paragraphs

# ----------- Tone Classification -----------

def classify_tone(paragraph, keyword, client):
    """
    Classifies the tone of a paragraph regarding a specific keyword.
    
    Uses OpenAI's GPT-4o model to analyze the sentiment and stance of a paragraph
    in relation to a given keyword, categorizing it as Supportive, Neutral, or Opposing.
    
    Args:
        paragraph (str): Text paragraph to analyze for tone
        keyword (str): Target keyword to analyze tone for
        client (OpenAI): OpenAI client instance for API calls
    
    Returns:
        str: Tone classification with explanation in format:
             "Supportive/Neutral/Opposing: explanation text"
    
    Raises:
        Exception: If OpenAI API call fails
    
    Example:
        >>> paragraph = "VR technology reduces motion sickness significantly."
        >>> keyword = "motion sickness"
        >>> result = classify_tone(paragraph, keyword, client)
        >>> print(result)
        "Supportive: The paragraph indicates VR reduces motion sickness."
    """
    prompt = (
        f"Classify the tone of the following paragraph with respect to the keyword: {keyword}\n\n"
        "There are three possible tones:\n"
        "1. Supportive: The paragraph affirms or supports the presence of an effect, relationship, or influence of the keyword.\n"
        "2. Neutral: The paragraph mentions the keyword without taking a stance.\n"
        "3. Opposing: The paragraph indicates no effect or contradictory evidence.\n\n"
        f"Paragraph: {paragraph}\n\n"
        "Respond with: Supportive, Neutral, or Opposing, and explain why."
    )
    response = client.chat.completions.create(
        model="gpt-4o",
        messages=[
            {"role": "system", "content": "You are a helpful academic tone analysis assistant."},
            {"role": "user", "content": prompt}
        ],
        temperature=0
    )
    return response.choices[0].message.content.strip()

# ----------- Main Execution -----------

def run_analysis(pdf_path, keywords, output_path, status_text):
    """
    Main analysis pipeline that processes PDF and generates results.
    
    Orchestrates the complete analysis workflow: extracts text from PDF,
    segments into paragraphs, analyzes tone for each keyword, calculates
    scores, and generates formatted Excel output.
    
    Args:
        pdf_path (str): Absolute path to the input PDF file
        keywords (list): List of keywords to analyze in the document
        output_path (str): Path where to save the output Excel file
        status_text (ScrolledText): Widget to display status updates
    
    Returns:
        None: Results are saved directly to the specified output path
    
    Raises:
        Exception: If PDF cannot be opened or processed
        Exception: If OpenAI API calls fail
        Exception: If Excel file cannot be created
    
    Process Flow:
        1. Extract text from all PDF pages
        2. Split text into paragraphs using AI
        3. Filter paragraphs containing each keyword
        4. Classify tone for each relevant paragraph
        5. Calculate scores and cumulative metrics
        6. Generate formatted Excel output with styling
    
    Example:
        >>> keywords = ["cybersickness", "VR", "motion sickness"]
        >>> run_analysis("research_paper.pdf", keywords, "results.xlsx", status_widget)
        # Creates results.xlsx with analysis
    """
    def update_status(message):
        """Helper function to update status in GUI"""
        status_text.insert(tk.END, message + "\n")
        status_text.see(tk.END)
        status_text.update()
    
    update_status("Starting analysis...")
    doc = fitz.open(pdf_path)
    all_paragraphs = []
    for page in doc:
        page_text = page.get_text("text")
        paragraphs = extract_paragraphs(page_text, client)
        all_paragraphs.extend(paragraphs)
    update_status(f"Extracted {len(all_paragraphs)} paragraphs.")

    tone_score = {"Supportive": 1, "Neutral": 0, "Opposing": -1}
    excel_rows = []
    overall_tones = []
    row_ranges = []

    for keyword in keywords:
        update_status(f"Processing keyword: {keyword}")
        keyword_paragraphs = [p for p in all_paragraphs if keyword.lower() in p.lower()]
        scores = []
        start_row = len(excel_rows) + 1  # +1 for header (Excel is 1-based)
        if not keyword_paragraphs:
            # Add a "not discussed" row
            excel_rows.append([
                keyword.capitalize(),
                "-",
                f"This paper does not discuss '{keyword}'.",
                "Not Discussed",
                "The keyword was not found in the document.",
                0,
                0,
                "Not Discussed"
            ])
            end_row = len(excel_rows)  # Only one row added
            row_ranges.append((start_row, end_row))
            overall_tones.append("Not Discussed")
            continue
        for i, para in enumerate(keyword_paragraphs):
            update_status(f"Analyzing paragraph {i+1} of {len(keyword_paragraphs)} for '{keyword}'")
            tone_full = classify_tone(para, keyword, client)
            if "." in tone_full:
                tone, explanation = tone_full.split(".", 1)
                tone = tone.strip()
                explanation = explanation.strip()
            else:
                tone = tone_full.strip()
                explanation = ""
            score = tone_score.get(tone, 0)
            scores.append(score)
            cumulative = sum(scores)
            excel_rows.append([
                keyword.capitalize(),
                f"Paragraph {i+1}",
                para,           # Extracted Text
                tone,
                explanation,    # New Explanation column
                score,
                cumulative,
                ""  # Placeholder for Overall Tone
            ])
        # Calculate overall tone for this keyword
        avg_score = sum(scores) / len(scores) if scores else 0
        if avg_score > 0:
            overall = "Supportive"
        elif avg_score < 0:
            overall = "Opposing"
        else:
            overall = "Neutral"
        # Store the range of rows for this keyword to merge later
        end_row = len(excel_rows)  # End row is the last row for this keyword
        row_ranges.append((start_row, end_row))
        overall_tones.append(overall)

    # Fill in the Overall Tone column for each keyword's rows
    for (start, end), overall in zip(row_ranges, overall_tones):
        for i in range(start-1, end):  # start-1 to end-1 inclusive
            excel_rows[i][7] = overall  # 7 is the 8th column (Overall Tone)

    df = pd.DataFrame(
        excel_rows,
        columns=[
            "Keyword",
            "Paragraph #",
            "Extracted Text",
            "Tone",
            "Explanation",
            "Score",
            "Cumulative Score",
            "Overall Tone"
        ]
    )
    
    if os.path.exists(output_path):
        try:
            os.remove(output_path)
            update_status("Removed existing output file.")
        except PermissionError:
            update_status("Excel file is open. Cannot overwrite.")
            messagebox.showerror(
                "Error",
                f"Cannot overwrite '{output_path}'.\nPlease close the file in Excel and try again."
            )
            return
    update_status("Saving results to Excel...")
    df.to_excel(output_path, index=False)
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    # Merge cells ONLY for Overall Tone column (column 8), not for Keyword
    for (start, end) in row_ranges:
        if end - start > 1:  # Only merge if more than one row for the keyword
            ws.merge_cells(start_row=start+1, start_column=8, end_row=end+1, end_column=8)  # Overall Tone

    # --- Formatting for cleaner Excel output ---
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)  # Limit width

    # Wrap text for Extracted Text and Explanation columns
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=5):  # Columns C, D, E
        for cell in row:
            cell.alignment = wrap_alignment

    # Header formatting
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add borders to all cells
    thin = Side(border_style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border

    wb.save(output_path)
    update_status("Excel file saved successfully!")
    update_status(f"Analysis complete! Results saved to: {output_path}")

# --- GUI Functions ---

def browse_pdf():
    """Opens file dialog for PDF selection using askopenfilename()"""
    filename = filedialog.askopenfilename(
        title="Select PDF file to analyze",
        filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
    )
    if filename:
        pdf_path_var.set(filename)

def browse_save_location():
    """Opens file dialog for save location selection using asksaveasfilename()"""
    filename = filedialog.asksaveasfilename(
        title="Save Analysis Results As",
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        initialfile="qualitative_analysis_results.xlsx"
    )
    if filename:
        output_path_var.set(filename)

def start_analysis():
    """Initiates the analysis process from GUI inputs with validation and error handling."""
    pdf_path = pdf_path_var.get()
    keywords_text = keywords_var.get()
    output_path = output_path_var.get()
    
    # Clear status text
    status_text.delete(1.0, tk.END)
    
    # Validation
    if not pdf_path:
        messagebox.showerror("Error", "Please select a PDF file.")
        return
    
    # Get keywords list
    keywords = [k.strip() for k in keywords_text.split(",") if k.strip()]
    
    if not keywords:
        messagebox.showerror("Error", "Please enter at least one keyword.")
        return
    
    # Check if output path is provided
    if not output_path.strip():
        messagebox.showerror("Error", "Please select where to save the output file.")
        return
    
    status_text.insert(tk.END, f"Will save results to: {output_path}\n\n")
    
    try:
        run_analysis(pdf_path, keywords, output_path, status_text)
        messagebox.showinfo("Analysis Complete", f"Analysis finished successfully!\nResults saved to: {output_path}")
        # Auto-close the GUI window after showing completion message
        root.quit()
        root.destroy()
    except Exception as e:
        error_msg = f"An error occurred during analysis:\n{str(e)}"
        status_text.insert(tk.END, f"ERROR: {str(e)}\n")
        status_text.see(tk.END)
        messagebox.showerror("Error", error_msg)

# --- GUI Setup ---

root = tk.Tk()
root.title("Qualitative Data Analyzer")
root.geometry("800x600")

# Variables for GUI inputs
pdf_path_var = tk.StringVar()
keywords_var = tk.StringVar()
output_path_var = tk.StringVar()

# Row 0: PDF File Selection
tk.Label(root, text="PDF File:", font=("Arial", 10, "bold")).grid(row=0, column=0, sticky="w", padx=5, pady=5)
pdf_entry = tk.Entry(root, textvariable=pdf_path_var, width=60)
pdf_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
tk.Button(root, text="Browse PDF", command=browse_pdf, width=12).grid(row=0, column=2, padx=5, pady=5)

# Row 1: Keywords Entry
tk.Label(root, text="Keywords (comma separated):", font=("Arial", 10, "bold")).grid(row=1, column=0, sticky="w", padx=5, pady=5)
keywords_entry = tk.Entry(root, textvariable=keywords_var, width=60)
keywords_entry.grid(row=1, column=1, columnspan=2, padx=5, pady=5, sticky="ew")

# Row 2: Output File Selection
tk.Label(root, text="Save To:", font=("Arial", 10, "bold")).grid(row=2, column=0, sticky="w", padx=5, pady=5)
output_entry = tk.Entry(root, textvariable=output_path_var, width=60)
output_entry.grid(row=2, column=1, padx=5, pady=5, sticky="ew")
tk.Button(root, text="Select File Location", command=browse_save_location, width=16).grid(row=2, column=2, padx=5, pady=5)

# Row 3: Run Analysis Button
tk.Button(root, text="Run Analysis", command=start_analysis, 
          font=("Arial", 12, "bold"), bg="#4CAF50", fg="white", 
          width=15, height=2).grid(row=3, column=1, pady=20)

# Row 4: Status Text (ScrolledText widget)
tk.Label(root, text="Status:", font=("Arial", 10, "bold")).grid(row=4, column=0, sticky="nw", padx=5, pady=5)
status_text = scrolledtext.ScrolledText(root, height=15, width=80, wrap=tk.WORD)
status_text.grid(row=4, column=1, columnspan=2, padx=5, pady=5, sticky="nsew")

# Configure grid weights for responsive layout
root.grid_columnconfigure(1, weight=1)
root.grid_rowconfigure(4, weight=1)

# Add initial message to status
status_text.insert(tk.END, "Ready to analyze PDF files.\nSelect a PDF file and enter keywords to begin.\nThe Excel file will be created automatically when you run the analysis.\n\n")

root.mainloop()

